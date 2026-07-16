from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, Iterable

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError


# ============================================================
# 基本設定
# ============================================================

APP_TITLE = "楽天銀行 クリエイティブ一次チェック"
SUPPORTED_IMAGE_EXTENSIONS = {"jpg", "jpeg", "png"}

REQUIRED_MASTER_COLUMNS: dict[str, list[str]] = {
    "01_NGワード": [
        "rule_id", "ステータス", "カテゴリ", "対象", "NG表現", "一致方法",
        "判定結果", "NG理由として出力する文言",
        "AI確認要否", "AI確認カテゴリ", "AIへの確認事項", "AI確認優先度",
    ],
    "02_必須文言": [
        "rule_id", "ステータス", "カテゴリ", "対象", "適用条件コード",
        "必須文言", "一致方法", "判定結果", "未検出時の出力文言",
        "AI確認要否", "AI確認カテゴリ", "AIへの確認事項", "AI確認優先度",
    ],
    "03_注意ワード": [
        "rule_id", "ステータス", "カテゴリ", "対象", "注意ワード",
        "一致方法", "判定結果", "確認内容", "出力文言",
        "AI確認要否", "AI確認カテゴリ", "AIへの確認事項", "AI確認優先度",
    ],
    "04_訴求別_必須注釈": [
        "rule_id", "ステータス", "カテゴリ", "対象", "検出ワード",
        "一致方法", "必須注釈", "注釈一致方法", "判定結果",
        "未検出時の出力文言", "AI確認要否", "AI確認カテゴリ",
        "AIへの確認事項", "AI確認優先度",
    ],
    "05_媒体_形式条件": [
        "rule_id", "ステータス", "媒体/用途", "媒体種別", "対象項目",
        "許可形式/基準値", "大小区分", "判定結果",
        "NG理由として出力する文言", "AI確認要否",
        "AI確認カテゴリ", "AIへの確認事項", "AI確認優先度",
    ],
    "06_AI確認項目": [
        "ai_check_id", "ステータス", "対象", "カテゴリ", "適用条件",
        "確認事項", "優先度", "回答必須", "備考",
    ],
}

PRIORITY_ORDER = {"高": 1, "中": 2, "低": 3}
JUDGMENT_ORDER = {"NG": 1, "要確認": 2, "OK": 3}

HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NG_FILL = PatternFill("solid", fgColor="F4CCCC")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")


# ============================================================
# データ型
# ============================================================

@dataclass
class CheckResult:
    file_name: str
    relative_path: str
    rule_id: str
    rule_type: str
    category: str
    judgment: str
    matched_text: str
    message: str
    check_detail: str
    ai_check_required: str
    ai_check_category: str
    ai_check_question: str
    ai_priority: str
    detected_by: str = "Python"


@dataclass
class FileSummary:
    file_name: str
    relative_path: str
    extension: str
    width_px: int | None
    height_px: int | None
    aspect_ratio: str
    color_mode: str
    file_size_mb: float
    sha256: str
    ocr_text: str
    overall_judgment: str
    ng_count: int
    review_count: int


@dataclass
class AICheckItem:
    file_name: str
    relative_path: str
    source: str
    rule_id: str
    category: str
    priority: str
    question: str
    answer_required: str


# ============================================================
# 共通関数
# ============================================================

def clean_text(value: Any) -> str:
    """Excel由来のNaNを空文字へ変換し、文字列を整える。"""
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalized_text(value: str) -> str:
    """
    OCR照合用の正規化。
    空白・改行・全角スペースを除去し、英字は小文字化する。
    """
    return re.sub(r"[\s\u3000]+", "", clean_text(value)).lower()


def safe_file_stem(file_name: str) -> str:
    stem = PurePosixPath(file_name.replace("\\", "/")).stem
    return re.sub(r'[\\/:*?"<>|]+', "_", stem).strip() or "creative"


def relative_upload_path(uploaded_file: Any) -> str:
    """
    フォルダアップロードではUploadedFile.nameに相対パスが入る場合がある。
    絶対パス化や親ディレクトリ参照を避け、安全な相対表記にする。
    """
    raw_name = clean_text(getattr(uploaded_file, "name", "uploaded_image"))
    parts = [
        p for p in raw_name.replace("\\", "/").split("/")
        if p not in {"", ".", ".."}
    ]
    return "/".join(parts) if parts else "uploaded_image"


def extension_of(file_name: str) -> str:
    suffix = PurePosixPath(file_name).suffix.lower().lstrip(".")
    return suffix


def is_active(row: pd.Series) -> bool:
    return clean_text(row.get("ステータス")) == "有効"


def matches_rule(text: str, pattern: str, match_method: str) -> bool:
    """マスタの一致方法に応じてOCRテキストを照合する。"""
    source = normalized_text(text)
    target = normalized_text(pattern)

    if not target:
        return False

    method = clean_text(match_method)

    if method == "完全一致":
        return source == target

    if method == "正規表現":
        try:
            return re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE) is not None
        except re.error:
            return False

    # マスタの標準は部分一致
    return target in source


def overall_judgment(results: list[CheckResult]) -> str:
    judgments = {r.judgment for r in results}
    if "NG" in judgments:
        return "NG"
    if "要確認" in judgments:
        return "要確認"
    return "OK"


def sort_results(results: list[CheckResult]) -> list[CheckResult]:
    return sorted(
        results,
        key=lambda x: (
            JUDGMENT_ORDER.get(x.judgment, 99),
            PRIORITY_ORDER.get(x.ai_priority, 99),
            x.rule_id,
        ),
    )


# ============================================================
# マスタ読み込み・検証
# ============================================================

@st.cache_data(show_spinner=False)
def load_master_from_bytes(master_bytes: bytes) -> dict[str, pd.DataFrame]:
    xls = pd.ExcelFile(BytesIO(master_bytes), engine="openpyxl")
    missing_sheets = [
        sheet_name
        for sheet_name in REQUIRED_MASTER_COLUMNS
        if sheet_name not in xls.sheet_names
    ]
    if missing_sheets:
        raise ValueError(
            "必要なシートがありません："
            + "、".join(missing_sheets)
        )

    master: dict[str, pd.DataFrame] = {}
    errors: list[str] = []

    for sheet_name, required_columns in REQUIRED_MASTER_COLUMNS.items():
        df = pd.read_excel(
            BytesIO(master_bytes),
            sheet_name=sheet_name,
            dtype=str,
            engine="openpyxl",
        )
        df.columns = [clean_text(col) for col in df.columns]
        df = df.dropna(how="all").fillna("")

        missing_columns = [
            col for col in required_columns if col not in df.columns
        ]
        if missing_columns:
            errors.append(
                f"{sheet_name}：不足列 "
                + "、".join(missing_columns)
            )

        id_column = "ai_check_id" if sheet_name == "06_AI確認項目" else "rule_id"
        if id_column in df.columns:
            ids = df[id_column].map(clean_text)
            duplicated = ids[(ids != "") & ids.duplicated(keep=False)].unique().tolist()
            if duplicated:
                errors.append(
                    f"{sheet_name}：ID重複 "
                    + "、".join(map(str, duplicated))
                )

        master[sheet_name] = df

    if errors:
        raise ValueError("\n".join(errors))

    return master


# ============================================================
# OCR
# ============================================================

def get_ocr_package_status() -> tuple[bool, str]:
    """OCRパッケージが環境に導入されているかを実行前に確認する。"""
    try:
        import rapidocr  # noqa: F401
        return True, "rapidocr"
    except ImportError:
        pass

    try:
        import rapidocr_onnxruntime  # noqa: F401
        return True, "rapidocr_onnxruntime"
    except ImportError:
        return False, ""


@st.cache_resource(show_spinner=False)
def get_ocr_engine():
    """
    RapidOCRを遅延ロードする。
    現行の rapidocr を優先し、旧 rapidocr_onnxruntime にも対応する。
    """
    try:
        from rapidocr import RapidOCR
        return RapidOCR()
    except ImportError:
        pass

    try:
        from rapidocr_onnxruntime import RapidOCR
        return RapidOCR()
    except ImportError as exc:
        raise RuntimeError(
            "OCRライブラリがインストールされていません。"
            "requirements.txtに rapidocr を追加し、Streamlitアプリを再起動してください。"
        ) from exc


@st.cache_data(show_spinner=False)
def run_ocr(image_bytes: bytes) -> str:
    image = Image.open(BytesIO(image_bytes)).convert("RGB")
    engine = get_ocr_engine()
    output = engine(image)

    # rapidocr v3はRapidOCROutput、旧版は(result, elapsed)を返す。
    if hasattr(output, "txts"):
        return "\n".join(
            clean_text(value)
            for value in (output.txts or [])
            if clean_text(value)
        )

    if isinstance(output, tuple):
        result = output[0]
    else:
        result = output

    if not result:
        return ""

    texts: list[str] = []
    for item in result:
        if len(item) >= 2:
            value = clean_text(item[1])
            if value:
                texts.append(value)

    return "\n".join(texts)


# ============================================================
# 画像情報
# ============================================================

def inspect_image(image_bytes: bytes) -> dict[str, Any]:
    try:
        image = Image.open(BytesIO(image_bytes))
        image.verify()

        image = Image.open(BytesIO(image_bytes))
        width, height = image.size
        ratio = f"{width / height:.4f}" if height else ""
        return {
            "width_px": width,
            "height_px": height,
            "aspect_ratio": ratio,
            "color_mode": clean_text(image.mode),
        }
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ValueError("画像ファイルを正常に読み込めません。") from exc


# ============================================================
# Pythonルール判定
# ============================================================

def make_result(
    file_name: str,
    relative_path: str,
    row: pd.Series,
    rule_type: str,
    matched_text: str,
    message: str,
    check_detail: str = "",
) -> CheckResult:
    return CheckResult(
        file_name=file_name,
        relative_path=relative_path,
        rule_id=clean_text(row.get("rule_id")),
        rule_type=rule_type,
        category=clean_text(row.get("カテゴリ")),
        judgment=clean_text(row.get("判定結果")) or "要確認",
        matched_text=matched_text,
        message=message,
        check_detail=check_detail,
        ai_check_required=clean_text(row.get("AI確認要否")),
        ai_check_category=clean_text(row.get("AI確認カテゴリ")),
        ai_check_question=clean_text(row.get("AIへの確認事項")),
        ai_priority=clean_text(row.get("AI確認優先度")) or "中",
    )


def check_format_rules(
    file_name: str,
    relative_path: str,
    file_size_mb: float,
    width_px: int,
    height_px: int,
    master_df: pd.DataFrame,
) -> list[CheckResult]:
    results: list[CheckResult] = []
    extension = extension_of(file_name)

    actual_values: dict[str, Any] = {
        "拡張子": extension,
        "横幅px": width_px,
        "縦幅px": height_px,
        "ファイルサイズMB": file_size_mb,
    }

    for _, row in master_df.iterrows():
        if not is_active(row):
            continue
        if clean_text(row.get("媒体種別")) != "画像":
            continue

        target_item = clean_text(row.get("対象項目"))
        standard = clean_text(row.get("許可形式/基準値"))
        comparison = clean_text(row.get("大小区分"))
        actual = actual_values.get(target_item)

        # 基準値が未設定なら判定しない
        if actual is None or not standard:
            continue

        violated = False

        if target_item == "拡張子":
            allowed = {
                item.strip().lower().lstrip(".")
                for item in standard.split(",")
                if item.strip()
            }
            violated = extension not in allowed
        else:
            try:
                standard_num = float(standard)
                actual_num = float(actual)
            except (TypeError, ValueError):
                continue

            if comparison == "以下":
                violated = actual_num > standard_num
            elif comparison == "以上":
                violated = actual_num < standard_num
            else:
                violated = actual_num != standard_num

        if violated:
            results.append(
                make_result(
                    file_name=file_name,
                    relative_path=relative_path,
                    row=row,
                    rule_type="媒体・形式条件",
                    matched_text=f"{target_item}: {actual}",
                    message=clean_text(row.get("NG理由として出力する文言")),
                    check_detail=f"基準値：{standard}",
                )
            )

    return results


def check_ng_words(
    file_name: str,
    relative_path: str,
    ocr_text: str,
    master_df: pd.DataFrame,
) -> list[CheckResult]:
    results: list[CheckResult] = []

    for _, row in master_df.iterrows():
        if not is_active(row):
            continue

        expression = clean_text(row.get("NG表現"))
        if matches_rule(ocr_text, expression, clean_text(row.get("一致方法"))):
            results.append(
                make_result(
                    file_name,
                    relative_path,
                    row,
                    "NGワード",
                    expression,
                    clean_text(row.get("NG理由として出力する文言")),
                    f"推奨表現：{clean_text(row.get('OK/推奨表現'))}",
                )
            )

    return results


def condition_is_applicable(condition_code: str, selected_conditions: set[str]) -> bool:
    code = clean_text(condition_code).upper()
    if not code or code == "ALL":
        return True
    return code in selected_conditions


def check_required_words(
    file_name: str,
    relative_path: str,
    ocr_text: str,
    selected_conditions: set[str],
    master_df: pd.DataFrame,
) -> list[CheckResult]:
    results: list[CheckResult] = []

    for _, row in master_df.iterrows():
        if not is_active(row):
            continue

        condition_code = clean_text(row.get("適用条件コード")).upper()
        if not condition_is_applicable(condition_code, selected_conditions):
            continue

        target = clean_text(row.get("対象"))
        required_text = clean_text(row.get("必須文言"))

        # 人間確認対象はOCR検出の有無にかかわらず出力
        if target == "人間確認":
            results.append(
                make_result(
                    file_name,
                    relative_path,
                    row,
                    "必須文言・設定",
                    required_text,
                    clean_text(row.get("未検出時の出力文言")),
                    clean_text(row.get("適用条件メモ")),
                )
            )
            continue

        found = matches_rule(
            ocr_text,
            required_text,
            clean_text(row.get("一致方法")),
        )
        if not found:
            results.append(
                make_result(
                    file_name,
                    relative_path,
                    row,
                    "必須文言",
                    required_text,
                    clean_text(row.get("未検出時の出力文言")),
                    clean_text(row.get("適用条件メモ")),
                )
            )

    return results


def check_warning_words(
    file_name: str,
    relative_path: str,
    ocr_text: str,
    master_df: pd.DataFrame,
) -> list[CheckResult]:
    results: list[CheckResult] = []

    for _, row in master_df.iterrows():
        if not is_active(row):
            continue

        warning_word = clean_text(row.get("注意ワード"))
        if matches_rule(
            ocr_text,
            warning_word,
            clean_text(row.get("一致方法")),
        ):
            results.append(
                make_result(
                    file_name,
                    relative_path,
                    row,
                    "注意ワード",
                    warning_word,
                    clean_text(row.get("出力文言")),
                    clean_text(row.get("確認内容")),
                )
            )

    return results


def check_required_annotations(
    file_name: str,
    relative_path: str,
    ocr_text: str,
    master_df: pd.DataFrame,
) -> list[CheckResult]:
    results: list[CheckResult] = []

    for _, row in master_df.iterrows():
        if not is_active(row):
            continue

        trigger_word = clean_text(row.get("検出ワード"))
        trigger_found = matches_rule(
            ocr_text,
            trigger_word,
            clean_text(row.get("一致方法")),
        )
        if not trigger_found:
            continue

        annotation = clean_text(row.get("必須注釈"))
        annotation_found = matches_rule(
            ocr_text,
            annotation,
            clean_text(row.get("注釈一致方法")),
        )

        if not annotation_found:
            results.append(
                make_result(
                    file_name,
                    relative_path,
                    row,
                    "訴求別必須注釈",
                    trigger_word,
                    clean_text(row.get("未検出時の出力文言")),
                    f"必要な注釈：{annotation}",
                )
            )

    return results


def run_python_checks(
    file_name: str,
    relative_path: str,
    image_bytes: bytes,
    image_info: dict[str, Any],
    ocr_text: str,
    selected_conditions: set[str],
    master: dict[str, pd.DataFrame],
) -> list[CheckResult]:
    file_size_mb = len(image_bytes) / (1024 * 1024)

    results: list[CheckResult] = []
    results.extend(
        check_format_rules(
            file_name,
            relative_path,
            file_size_mb,
            int(image_info["width_px"]),
            int(image_info["height_px"]),
            master["05_媒体_形式条件"],
        )
    )
    results.extend(
        check_ng_words(
            file_name,
            relative_path,
            ocr_text,
            master["01_NGワード"],
        )
    )
    results.extend(
        check_required_words(
            file_name,
            relative_path,
            ocr_text,
            selected_conditions,
            master["02_必須文言"],
        )
    )
    results.extend(
        check_warning_words(
            file_name,
            relative_path,
            ocr_text,
            master["03_注意ワード"],
        )
    )
    results.extend(
        check_required_annotations(
            file_name,
            relative_path,
            ocr_text,
            master["04_訴求別_必須注釈"],
        )
    )

    return sort_results(results)


# ============================================================
# AI確認事項・プロンプト
# ============================================================

def ai_target_applies(target: str, media_type: str = "画像") -> bool:
    normalized = clean_text(target)
    return normalized in {"", "ALL", media_type, "画像・動画"}


def ai_condition_applies(condition: str, selected_conditions: set[str]) -> bool:
    condition = clean_text(condition).upper()
    if condition in {"", "ALL"}:
        return True
    return condition in selected_conditions


def remove_duplicate_ai_items(items: list[AICheckItem]) -> list[AICheckItem]:
    unique: list[AICheckItem] = []
    seen: set[str] = set()

    for item in items:
        key = normalized_text(item.question)
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(item)

    return unique


def build_ai_check_items(
    file_name: str,
    relative_path: str,
    python_results: list[CheckResult],
    common_ai_df: pd.DataFrame,
    selected_conditions: set[str],
) -> list[AICheckItem]:
    items: list[AICheckItem] = []

    # ルール検出に関係なく行う共通AI確認
    for _, row in common_ai_df.iterrows():
        if not is_active(row):
            continue
        if not ai_target_applies(clean_text(row.get("対象")), "画像"):
            continue
        if not ai_condition_applies(
            clean_text(row.get("適用条件")),
            selected_conditions,
        ):
            continue

        items.append(
            AICheckItem(
                file_name=file_name,
                relative_path=relative_path,
                source="共通AI確認",
                rule_id=clean_text(row.get("ai_check_id")),
                category=clean_text(row.get("カテゴリ")),
                priority=clean_text(row.get("優先度")) or "中",
                question=clean_text(row.get("確認事項")),
                answer_required=clean_text(row.get("回答必須")),
            )
        )

    # Python検出ルールに紐づく重点確認
    for result in python_results:
        if result.ai_check_required not in {"必要", "常時"}:
            continue
        if not result.ai_check_question:
            continue

        items.append(
            AICheckItem(
                file_name=file_name,
                relative_path=relative_path,
                source="Python検出結果",
                rule_id=result.rule_id,
                category=result.ai_check_category or result.category,
                priority=result.ai_priority or "中",
                question=result.ai_check_question,
                answer_required="はい",
            )
        )

    items = remove_duplicate_ai_items(items)
    return sorted(
        items,
        key=lambda x: (
            PRIORITY_ORDER.get(x.priority, 99),
            x.category,
            x.rule_id,
        ),
    )


def build_ai_prompt(
    summary: FileSummary,
    python_results: list[CheckResult],
    ai_items: list[AICheckItem],
    selected_condition_labels: list[str],
) -> str:
    if python_results:
        result_lines = [
            (
                f"- [{r.judgment}] {r.rule_id}／{r.category}\n"
                f"  検出・不足内容：{r.matched_text or '―'}\n"
                f"  出力内容：{r.message}\n"
                f"  確認補足：{r.check_detail or '―'}"
            )
            for r in python_results
        ]
        python_result_text = "\n".join(result_lines)
    else:
        python_result_text = "明確なNG・要確認項目は検出されませんでした。"

    if ai_items:
        ai_check_text = "\n".join(
            f"{index}. [{item.priority}] [{item.category}] {item.question}"
            for index, item in enumerate(ai_items, start=1)
        )
    else:
        ai_check_text = "クリエイティブ全体の視認性と誤認リスクを確認してください。"

    condition_text = (
        "、".join(selected_condition_labels)
        if selected_condition_labels
        else "共通条件のみ"
    )

    return f"""添付された広告クリエイティブを確認してください。

この確認は掲載可否を決定する最終審査ではありません。
Pythonによる機械判定結果を参考に、人間が確認すべき箇所を抽出する一次チェックとして回答してください。

【対象ファイル】
ファイル名：{summary.file_name}
相対パス：{summary.relative_path}
画像サイズ：{summary.width_px} × {summary.height_px}px
適用条件：{condition_text}

【OCR抽出テキスト】
{summary.ocr_text or "OCRテキストを取得できませんでした。画像を直接確認してください。"}

【Pythonによる判定結果】
{python_result_text}

【重点確認事項】
{ai_check_text}

【確認時の注意】
- 法務上の適法性を断定しないでください。
- 掲載可能・掲載不可の最終判断は行わないでください。
- 画像から判断できない条件、根拠、設定、権利関係は推測しないでください。
- 不明確な場合は「問題なし」と推測せず、「要確認」としてください。
- 問題箇所は画像内の位置、該当文言、周辺要素をできるだけ具体的に示してください。
- OCR結果に誤認識があり得るため、必ず添付画像の表示内容を優先してください。
- 各確認事項について個別に回答してください。

【回答形式】
総合判定候補：問題なし／要確認／修正推奨

確認結果：
1.
- ルールID：
- 確認項目：
- 判定：問題なし／要確認／修正推奨／画像のみでは確認不可
- 該当箇所：
- 理由：
- 人間が確認すべき内容：

人間確認事項：
-
""".strip()


# ============================================================
# Excel・ZIP出力
# ============================================================

def dataframe_from_dataclasses(items: Iterable[Any], columns: list[str]) -> pd.DataFrame:
    rows = [asdict(item) for item in items]
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows)


def style_worksheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:1000]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        ws.column_dimensions[column_letter].width = max(10, min(max_length + 2, 55))


def apply_result_colors(ws, judgment_header: str) -> None:
    headers = {
        cell.value: cell.column
        for cell in ws[1]
    }
    judgment_col = headers.get(judgment_header)
    if not judgment_col:
        return

    for row_num in range(2, ws.max_row + 1):
        value = clean_text(ws.cell(row=row_num, column=judgment_col).value)
        if value == "NG":
            fill = NG_FILL
        elif value == "要確認":
            fill = WARN_FILL
        elif value == "OK":
            fill = OK_FILL
        else:
            continue

        ws.cell(row=row_num, column=judgment_col).fill = fill


def add_thumbnail_images(
    ws,
    summaries: list[FileSummary],
    image_bytes_map: dict[str, bytes],
) -> None:
    """
    01_ファイル一覧のA列へ縮小画像を貼り付ける。
    元画像の縦横比を維持し、最大140×90px程度に収める。
    """
    ws.column_dimensions["A"].width = 24

    for row_num, summary in enumerate(summaries, start=2):
        image_bytes = image_bytes_map.get(summary.relative_path)
        if not image_bytes:
            continue

        try:
            pil_image = Image.open(BytesIO(image_bytes)).convert("RGB")
            pil_image.thumbnail((140, 90))

            thumbnail_stream = BytesIO()
            pil_image.save(thumbnail_stream, format="PNG")
            thumbnail_stream.seek(0)

            excel_image = XLImage(thumbnail_stream)
            excel_image.width = pil_image.width
            excel_image.height = pil_image.height

            # save時までBytesIOを保持する
            excel_image._thumbnail_stream = thumbnail_stream
            ws.add_image(excel_image, f"A{row_num}")
            ws.row_dimensions[row_num].height = 72
        except Exception:
            # サムネイル生成失敗だけでExcel出力全体を止めない
            ws.cell(row=row_num, column=1, value="画像貼付失敗")


def append_dataframe_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=sheet_name)

    for col_num, column_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_num, value=column_name)

    for row_num, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_num, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            ws.cell(row=row_num, column=col_num, value=value)

    style_worksheet(ws)


def create_result_excel(
    summaries: list[FileSummary],
    all_results: list[CheckResult],
    all_ai_items: list[AICheckItem],
    prompts: dict[str, str],
    errors: list[dict[str, str]],
    image_bytes_map: dict[str, bytes],
) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary_columns = list(FileSummary.__annotations__.keys())
    result_columns = list(CheckResult.__annotations__.keys())
    ai_columns = list(AICheckItem.__annotations__.keys())

    summary_df = dataframe_from_dataclasses(summaries, summary_columns)
    summary_df.insert(0, "thumbnail", "")
    results_df = dataframe_from_dataclasses(all_results, result_columns)
    ai_df = dataframe_from_dataclasses(all_ai_items, ai_columns)

    prompts_df = pd.DataFrame(
        [
            {
                "relative_path": relative_path,
                "AI用プロンプト": prompt,
            }
            for relative_path, prompt in prompts.items()
        ],
        columns=["relative_path", "AI用プロンプト"],
    )

    errors_df = pd.DataFrame(
        errors,
        columns=["relative_path", "エラー内容"],
    )

    append_dataframe_sheet(wb, "01_ファイル一覧", summary_df)
    add_thumbnail_images(
        wb["01_ファイル一覧"],
        summaries,
        image_bytes_map,
    )
    append_dataframe_sheet(wb, "02_Python判定結果", results_df)
    append_dataframe_sheet(wb, "03_AI確認事項", ai_df)
    append_dataframe_sheet(wb, "04_AIプロンプト", prompts_df)
    append_dataframe_sheet(wb, "05_エラー", errors_df)

    apply_result_colors(wb["01_ファイル一覧"], "overall_judgment")
    apply_result_colors(wb["02_Python判定結果"], "judgment")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_prompt_zip(prompts: dict[str, str]) -> bytes:
    output = BytesIO()

    with zipfile.ZipFile(
        output,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as zip_file:
        used_names: set[str] = set()

        for relative_path, prompt in prompts.items():
            base_name = f"{safe_file_stem(relative_path)}_AI確認プロンプト.txt"
            candidate = base_name
            count = 2
            while candidate in used_names:
                candidate = (
                    f"{safe_file_stem(relative_path)}_{count}"
                    "_AI確認プロンプト.txt"
                )
                count += 1
            used_names.add(candidate)

            zip_file.writestr(
                candidate,
                prompt.encode("utf-8-sig"),
            )

    return output.getvalue()


# ============================================================
# 画面表示
# ============================================================

def render_result_card(
    summary: FileSummary,
    results: list[CheckResult],
    ai_items: list[AICheckItem],
    prompt: str,
    image_bytes: bytes,
) -> None:
    status_icon = {
        "NG": "🔴",
        "要確認": "🟡",
        "OK": "🟢",
    }.get(summary.overall_judgment, "⚪")

    with st.expander(
        f"{status_icon} {summary.relative_path}｜{summary.overall_judgment}",
        expanded=summary.overall_judgment != "OK",
    ):
        image_col, info_col = st.columns([1, 1])

        with image_col:
            st.image(image_bytes, caption=summary.relative_path)

        with info_col:
            st.write(f"**サイズ**：{summary.width_px} × {summary.height_px}px")
            st.write(f"**容量**：{summary.file_size_mb:.3f}MB")
            st.write(f"**カラーモード**：{summary.color_mode}")
            st.write(f"**NG**：{summary.ng_count}件")
            st.write(f"**要確認**：{summary.review_count}件")

        result_tab, ocr_tab, ai_tab, prompt_tab = st.tabs(
            ["Python判定", "OCR結果", "AI確認事項", "AIプロンプト"]
        )

        with result_tab:
            if results:
                result_df = pd.DataFrame(
                    [
                        {
                            "判定": r.judgment,
                            "ルールID": r.rule_id,
                            "カテゴリ": r.category,
                            "検出・不足内容": r.matched_text,
                            "結果": r.message,
                            "確認補足": r.check_detail,
                        }
                        for r in results
                    ]
                )
                st.dataframe(
                    result_df,
                    width="stretch",
                    hide_index=True,
                )
            else:
                st.success("マスタに基づく明確なNG・要確認項目はありませんでした。")

        with ocr_tab:
            st.text_area(
                "抽出テキスト",
                value=summary.ocr_text,
                height=260,
                key=f"ocr_{summary.sha256}",
            )

        with ai_tab:
            if ai_items:
                ai_df = pd.DataFrame(
                    [
                        {
                            "優先度": item.priority,
                            "ルールID": item.rule_id,
                            "カテゴリ": item.category,
                            "確認事項": item.question,
                            "回答必須": item.answer_required,
                        }
                        for item in ai_items
                    ]
                )
                st.dataframe(
                    ai_df,
                    width="stretch",
                    hide_index=True,
                )
            else:
                st.info("AI確認事項はありません。")

        with prompt_tab:
            st.code(prompt, language=None)
            st.download_button(
                "このAIプロンプトをダウンロード",
                data=prompt.encode("utf-8-sig"),
                file_name=f"{safe_file_stem(summary.file_name)}_AI確認プロンプト.txt",
                mime="text/plain",
                key=f"prompt_download_{summary.sha256}",
            )


# ============================================================
# Streamlitアプリ本体
# ============================================================

def main() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="🔎",
        layout="wide",
    )

    st.title(APP_TITLE)
    st.caption(
        "Pythonで明確なルールを判定し、AI確認事項と確認用プロンプトを生成します。"
        "最終的な掲載可否・法務・ブランド判断は人間が行ってください。"
    )

    with st.sidebar:
        st.header("1. 判定マスタ")

        master_file = st.file_uploader(
            "AI項目追加版のマスタExcel",
            type=["xlsx"],
            accept_multiple_files=False,
            key="master_uploader",
            help="マスタはアプリ内に保存せず、アップロードされた内容をメモリ上で読み込みます。",
        )

        st.header("2. 追加ルール（任意）")
        st.caption(
            "画像の掲載先や内容が分かる場合だけ選択してください。"
            "未選択でも共通ルールのチェックは実行されます。"
        )

        selected_platform = st.multiselect(
            "この画像を掲載する媒体",
            options=["X", "Instagram", "YouTube"],
            default=[],
            placeholder="必要な場合のみ選択",
            help=(
                "選択した媒体に応じて「PR」表記などの媒体別ルールを追加します。"
                "単なるバナー画像の確認なら未選択で構いません。"
            ),
        )

        benefit_claim = st.checkbox(
            "画像内で特典・ポイントを訴求している",
            value=False,
            help=(
                "ポイント付与や口座開設特典などを訴求する画像でオンにします。"
                "特典条件や詳細ページ誘導の必須チェックを追加します。"
            ),
        )

        selected_conditions: set[str] = set()
        selected_condition_labels: list[str] = []

        platform_code_map = {
            "X": "X",
            "Instagram": "INSTAGRAM",
            "YouTube": "YOUTUBE",
        }
        for platform in selected_platform:
            selected_conditions.add(platform_code_map[platform])
            selected_condition_labels.append(platform)

        if benefit_claim:
            selected_conditions.add("BENEFIT")
            selected_condition_labels.append("特典・ポイント記載あり")

        st.header("3. OCR")
        ocr_available, ocr_package_name = get_ocr_package_status()

        if ocr_available:
            st.success(f"OCR利用可能：{ocr_package_name}")
        else:
            st.error(
                "OCRライブラリが未導入です。"
                "リポジトリ直下のrequirements.txtを更新して、"
                "StreamlitアプリをRebootしてください。"
            )

        run_ocr_enabled = st.checkbox(
            "OCRを実行する",
            value=ocr_available,
            disabled=not ocr_available,
            help="RapidOCRで画像内の文字を抽出します。",
        )

    st.subheader("チェック対象画像")
    st.caption(
        "画像を複数選択するか、画像が入ったフォルダをこの欄へドラッグ＆ドロップしてください。"
    )

    selected_files = st.file_uploader(
        "画像ファイルまたはフォルダ",
        type=sorted(SUPPORTED_IMAGE_EXTENSIONS),
        accept_multiple_files=True,
        key="image_uploader",
    )

    uploaded_files: list[Any] = []
    seen_uploads: set[tuple[str, str]] = set()

    for uploaded in list(selected_files or []):
        relative_path = relative_upload_path(uploaded)
        file_bytes = uploaded.getvalue()
        content_hash = hashlib.sha256(file_bytes).hexdigest()
        key = (relative_path.lower(), content_hash)
        if key in seen_uploads:
            continue
        seen_uploads.add(key)
        uploaded_files.append(uploaded)

    if uploaded_files:
        st.info(f"チェック対象：{len(uploaded_files)}ファイル")

    run_button = st.button(
        "クリエイティブチェックを実行",
        type="primary",
        disabled=not (master_file and uploaded_files),
        width="stretch",
    )

    if not run_button:
        if not master_file:
            st.warning("最初に判定マスタExcelをアップロードしてください。")
        elif not uploaded_files:
            st.warning("画像ファイルまたは画像フォルダをアップロードしてください。")
        return

    try:
        with st.spinner("マスタを検証しています..."):
            master = load_master_from_bytes(master_file.getvalue())
    except Exception as exc:
        st.error(f"マスタを読み込めませんでした。\n\n{exc}")
        return

    st.success("マスタを正常に読み込みました。")

    summaries: list[FileSummary] = []
    all_results: list[CheckResult] = []
    all_ai_items: list[AICheckItem] = []
    prompts: dict[str, str] = {}
    errors: list[dict[str, str]] = []
    image_bytes_map: dict[str, bytes] = {}
    results_by_path: dict[str, list[CheckResult]] = {}
    ai_items_by_path: dict[str, list[AICheckItem]] = {}

    progress = st.progress(0, text="チェックを開始します。")
    total_files = len(uploaded_files)

    for index, uploaded in enumerate(uploaded_files, start=1):
        relative_path = relative_upload_path(uploaded)
        file_name = PurePosixPath(relative_path).name
        image_bytes = uploaded.getvalue()

        progress.progress(
            (index - 1) / total_files,
            text=f"{relative_path} を処理しています...",
        )

        try:
            ext = extension_of(file_name)
            if ext not in SUPPORTED_IMAGE_EXTENSIONS:
                raise ValueError(f"対応外の画像形式です：{ext or '拡張子なし'}")

            image_info = inspect_image(image_bytes)

            ocr_warning = ""
            if run_ocr_enabled:
                try:
                    ocr_text = run_ocr(image_bytes)
                except Exception as ocr_exc:
                    ocr_text = ""
                    ocr_warning = str(ocr_exc)
            else:
                ocr_text = ""

            python_results = run_python_checks(
                file_name=file_name,
                relative_path=relative_path,
                image_bytes=image_bytes,
                image_info=image_info,
                ocr_text=ocr_text,
                selected_conditions=selected_conditions,
                master=master,
            )

            if ocr_warning:
                python_results.append(
                    CheckResult(
                        file_name=file_name,
                        relative_path=relative_path,
                        rule_id="SYS_OCR",
                        rule_type="システム",
                        category="OCR",
                        judgment="要確認",
                        matched_text="OCR未実行",
                        message=(
                            "OCRを実行できなかったため、文字ルールは未判定です。"
                            "画像情報と共通AI確認事項は出力しています。"
                        ),
                        check_detail=ocr_warning,
                        ai_check_required="不要",
                        ai_check_category="OCR",
                        ai_check_question="",
                        ai_priority="高",
                    )
                )
                python_results = sort_results(python_results)

            ai_items = build_ai_check_items(
                file_name=file_name,
                relative_path=relative_path,
                python_results=python_results,
                common_ai_df=master["06_AI確認項目"],
                selected_conditions=selected_conditions,
            )

            judgment = overall_judgment(python_results)
            file_size_mb = len(image_bytes) / (1024 * 1024)
            sha256 = hashlib.sha256(image_bytes).hexdigest()

            summary = FileSummary(
                file_name=file_name,
                relative_path=relative_path,
                extension=ext,
                width_px=image_info["width_px"],
                height_px=image_info["height_px"],
                aspect_ratio=image_info["aspect_ratio"],
                color_mode=image_info["color_mode"],
                file_size_mb=round(file_size_mb, 4),
                sha256=sha256,
                ocr_text=ocr_text,
                overall_judgment=judgment,
                ng_count=sum(r.judgment == "NG" for r in python_results),
                review_count=sum(
                    r.judgment == "要確認" for r in python_results
                ),
            )

            prompt = build_ai_prompt(
                summary=summary,
                python_results=python_results,
                ai_items=ai_items,
                selected_condition_labels=selected_condition_labels,
            )

            summaries.append(summary)
            all_results.extend(python_results)
            all_ai_items.extend(ai_items)
            prompts[relative_path] = prompt
            image_bytes_map[relative_path] = image_bytes
            results_by_path[relative_path] = python_results
            ai_items_by_path[relative_path] = ai_items

        except Exception as exc:
            errors.append(
                {
                    "relative_path": relative_path,
                    "エラー内容": str(exc),
                }
            )

    progress.progress(1.0, text="チェックが完了しました。")

    if not summaries:
        st.error("正常に処理できた画像がありませんでした。")
        if errors:
            st.dataframe(pd.DataFrame(errors), width="stretch", hide_index=True)
        return

    st.divider()
    st.subheader("集計")

    total_ng = sum(summary.ng_count for summary in summaries)
    total_review = sum(summary.review_count for summary in summaries)
    ok_files = sum(summary.overall_judgment == "OK" for summary in summaries)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("処理済み", len(summaries))
    col2.metric("NG項目", total_ng)
    col3.metric("要確認項目", total_review)
    col4.metric("Python判定OK", ok_files)

    result_excel = create_result_excel(
        summaries=summaries,
        all_results=all_results,
        all_ai_items=all_ai_items,
        prompts=prompts,
        errors=errors,
        image_bytes_map=image_bytes_map,
    )
    prompt_zip = create_prompt_zip(prompts)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    download_col1, download_col2 = st.columns(2)
    with download_col1:
        st.download_button(
            "結果Excelをダウンロード",
            data=result_excel,
            file_name=f"creative_check_result_{timestamp}.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            width="stretch",
        )

    with download_col2:
        st.download_button(
            "AIプロンプト一式をZIPでダウンロード",
            data=prompt_zip,
            file_name=f"ai_prompts_{timestamp}.zip",
            mime="application/zip",
            width="stretch",
        )

    if errors:
        st.warning(f"{len(errors)}ファイルでエラーが発生しました。")
        st.dataframe(
            pd.DataFrame(errors),
            width="stretch",
            hide_index=True,
        )

    st.divider()
    st.subheader("ファイル別結果")

    sorted_summaries = sorted(
        summaries,
        key=lambda x: (
            JUDGMENT_ORDER.get(x.overall_judgment, 99),
            x.relative_path.lower(),
        ),
    )

    for summary in sorted_summaries:
        path = summary.relative_path
        render_result_card(
            summary=summary,
            results=results_by_path[path],
            ai_items=ai_items_by_path[path],
            prompt=prompts[path],
            image_bytes=image_bytes_map[path],
        )


if __name__ == "__main__":
    main()
