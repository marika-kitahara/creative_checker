from __future__ import annotations

import zipfile
from dataclasses import asdict
from io import BytesIO
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from models import AICheckItem, CheckResult, FileSummary
from utils import clean_text, safe_file_stem


FONT_NAME = "Meiryo UI"
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
NG_FILL = PatternFill("solid", fgColor="F4CCCC")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")

SUMMARY_NAMES = {
    "thumbnail": "画像", "file_name": "ファイル名",
    "relative_path": "相対パス", "extension": "拡張子",
    "width_px": "横幅（px）", "height_px": "縦幅（px）",
    "aspect_ratio": "アスペクト比", "color_mode": "カラーモード",
    "file_size_mb": "ファイルサイズ（MB）", "sha256": "SHA-256",
    "ocr_text": "OCR抽出テキスト", "overall_judgment": "総合判定",
    "ng_count": "NG件数", "review_count": "要確認件数",
}

RESULT_NAMES = {
    "file_name": "ファイル名", "relative_path": "相対パス",
    "rule_id": "ルールID", "rule_type": "ルール種別",
    "category": "カテゴリ", "judgment": "判定",
    "matched_text": "検出・不足内容", "message": "判定結果・メッセージ",
    "check_detail": "確認補足", "ai_check_required": "AI確認要否",
    "ai_check_category": "AI確認カテゴリ",
    "ai_check_question": "AIへの確認事項",
    "ai_priority": "AI確認優先度", "detected_by": "検出方法",
}

AI_NAMES = {
    "file_name": "ファイル名", "relative_path": "相対パス",
    "source": "確認事項の出所", "rule_id": "ルールID",
    "category": "カテゴリ", "priority": "優先度",
    "question": "AI確認事項", "answer_required": "回答必須",
}


def _df(items: Iterable[Any], columns: list[str]) -> pd.DataFrame:
    rows = [asdict(item) for item in items]
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=columns)


def _style(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT_NAME, color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        length = 0
        for cell in cells[:1000]:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), 60))
        ws.column_dimensions[letter].width = max(10, min(length + 2, 55))


def _append(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    for col, column_name in enumerate(df.columns, 1):
        ws.cell(1, col, column_name)
    for row_num, row in enumerate(df.itertuples(index=False, name=None), 2):
        for col, value in enumerate(row, 1):
            ws.cell(row_num, col, "" if pd.isna(value) else value)
    _style(ws)


def _colors(ws, header: str) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    column = headers.get(header)
    if not column:
        return
    for row in range(2, ws.max_row + 1):
        value = clean_text(ws.cell(row, column).value)
        fill = {"NG": NG_FILL, "要確認": WARN_FILL, "OK": OK_FILL}.get(value)
        if fill:
            ws.cell(row, column).fill = fill


def _thumbnails(
    ws,
    summaries: list[FileSummary],
    image_bytes_map: dict[str, bytes],
) -> None:
    ws.column_dimensions["A"].width = 24
    for row_num, summary in enumerate(summaries, 2):
        data = image_bytes_map.get(summary.relative_path)
        if not data:
            continue
        try:
            image = Image.open(BytesIO(data)).convert("RGB")
            image.thumbnail((140, 90))
            stream = BytesIO()
            image.save(stream, format="PNG")
            stream.seek(0)
            xl_image = XLImage(stream)
            xl_image.width, xl_image.height = image.size
            xl_image._stream_ref = stream
            ws.add_image(xl_image, f"A{row_num}")
            ws.row_dimensions[row_num].height = 72
        except Exception:
            ws.cell(row_num, 1, "画像貼付失敗")


def create_result_excel(
    summaries: list[FileSummary],
    results: list[CheckResult],
    ai_items: list[AICheckItem],
    prompts: dict[str, str],
    errors: list[dict[str, str]],
    image_bytes_map: dict[str, bytes],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    summary_df = _df(summaries, list(FileSummary.__annotations__))
    summary_df.insert(0, "thumbnail", "")
    summary_df = summary_df.rename(columns=SUMMARY_NAMES)
    result_df = _df(results, list(CheckResult.__annotations__)).rename(
        columns=RESULT_NAMES
    )
    ai_df = _df(ai_items, list(AICheckItem.__annotations__)).rename(
        columns=AI_NAMES
    )
    prompt_df = pd.DataFrame(
        [{"相対パス": path, "AI用プロンプト": prompt}
         for path, prompt in prompts.items()],
        columns=["相対パス", "AI用プロンプト"],
    )
    error_df = pd.DataFrame(
        [{"相対パス": item.get("relative_path", ""),
          "エラー内容": item.get("エラー内容", "")}
         for item in errors],
        columns=["相対パス", "エラー内容"],
    )

    _append(wb, "01_ファイル一覧", summary_df)
    _thumbnails(wb["01_ファイル一覧"], summaries, image_bytes_map)
    _append(wb, "02_Python判定結果", result_df)
    _append(wb, "03_AI確認事項", ai_df)
    _append(wb, "04_AIプロンプト", prompt_df)
    _append(wb, "05_エラー", error_df)
    _colors(wb["01_ファイル一覧"], "総合判定")
    _colors(wb["02_Python判定結果"], "判定")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_prompt_zip(prompts: dict[str, str]) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        used = set()
        for path, prompt in prompts.items():
            base = f"{safe_file_stem(path)}_AI確認プロンプト.txt"
            name = base
            number = 2
            while name in used:
                name = f"{safe_file_stem(path)}_{number}_AI確認プロンプト.txt"
                number += 1
            used.add(name)
            archive.writestr(name, prompt.encode("utf-8-sig"))
    return output.getvalue()
